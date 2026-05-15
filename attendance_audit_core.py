from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
import re
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


RESULT_HEADERS = [
    "签到时间",
    "签退时间",
    "是否满足8小时",
    "打卡制式是否一致",
    "签到车牌",
    "签退车牌",
    "车牌是否一致",
    "备注",
]


@dataclass(frozen=True)
class AttendanceRecord:
    name: str
    email: str
    attendance_time: datetime
    product: str
    car_scene: str
    plate_number: str
    source_file: str = ""


@dataclass
class AuditResult:
    sign_in_time: datetime | None = None
    sign_out_time: datetime | None = None
    meets_8_hours: str = "否"
    product_consistency: str = "签到和签退制式不一致"
    sign_in_plate: str = ""
    sign_out_plate: str = ""
    plate_consistency: str = "缺少签到"
    remark: str = ""


@dataclass
class AuditSummary:
    person_count: int
    attendance_file_count: int
    attendance_record_count: int
    output_path: Path
    duplicate_names: list[str] = field(default_factory=list)
    merged_output_path: Path | None = None


class AttendanceAuditError(Exception):
    """Raised when the source workbooks cannot be audited."""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_attendance_time(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = normalize_text(value)
    if not text:
        return None

    date_time_matches = re.findall(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}\s+\d{1,2}:\d{2}(?::\d{2})?", text)
    if len(date_time_matches) > 1:
        text = date_time_matches[-1]

    formats = [
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise AttendanceAuditError(f"无法解析考勤时间：{text}")


def expected_product_keyword(person_format: str) -> str:
    text = normalize_text(person_format).upper()
    if "5G NR" in text or "5GNR" in text:
        return "5G NR"
    if "LTE" in text:
        return "LTE"
    return normalize_text(person_format)


def is_product_match(person_format: str, attendance_product: str) -> bool:
    expected = expected_product_keyword(person_format).upper()
    actual = normalize_text(attendance_product).upper()
    if not expected or not actual:
        return False
    if expected == "5G NR":
        return "5G NR" in actual or "5GNR" in actual
    if expected == "LTE":
        return "LTE" in actual
    return expected == actual


def judge_product_consistency(
    person_format: str,
    sign_in_record: AttendanceRecord | None,
    sign_out_record: AttendanceRecord | None,
) -> str:
    sign_in_ok = bool(sign_in_record and is_product_match(person_format, sign_in_record.product))
    sign_out_ok = bool(sign_out_record and is_product_match(person_format, sign_out_record.product))
    if sign_in_ok and sign_out_ok:
        return "一致"
    if not sign_in_ok and sign_out_ok:
        return "签到制式不一致"
    if sign_in_ok and not sign_out_ok:
        return "签退制式不一致"
    return "签到和签退制式不一致"


def judge_plate_consistency(sign_in_plate: str, sign_out_plate: str) -> str:
    sign_in_plate = normalize_text(sign_in_plate)
    sign_out_plate = normalize_text(sign_out_plate)
    if not sign_in_plate and not sign_out_plate:
        return "未用车"
    if not sign_in_plate:
        return "缺少签到"
    if not sign_out_plate:
        return "缺少签退"
    if sign_in_plate == sign_out_plate:
        return "一致"
    return "不一致"


def count_text(count: int) -> str:
    numbers = {
        2: "两",
        3: "三",
        4: "四",
        5: "五",
        6: "六",
        7: "七",
        8: "八",
        9: "九",
        10: "十",
    }
    return numbers.get(count, str(count))


def choose_plate_result(
    start_car_records: list[AttendanceRecord],
    end_car_records: list[AttendanceRecord],
) -> tuple[str, str, str]:
    start_records_with_plate = [record for record in start_car_records if normalize_text(record.plate_number)]
    end_records_with_plate = [record for record in end_car_records if normalize_text(record.plate_number)]
    if not start_records_with_plate and not end_records_with_plate:
        return "", "", "未用车"
    if not start_records_with_plate:
        sign_out_plate = end_records_with_plate[-1].plate_number if end_records_with_plate else ""
        return "", sign_out_plate, "缺少签到"
    if not end_records_with_plate:
        sign_in_plate = start_records_with_plate[0].plate_number
        return sign_in_plate, "", "缺少签退"

    end_plates = {normalize_text(record.plate_number) for record in end_records_with_plate}
    for start_record in start_records_with_plate:
        start_plate = normalize_text(start_record.plate_number)
        if start_plate in end_plates:
            return start_plate, start_plate, "一致"

    return start_records_with_plate[0].plate_number, end_records_with_plate[-1].plate_number, "不一致"


def audit_person_records(person_format: str, records: Iterable[AttendanceRecord]) -> AuditResult:
    records = list(records)
    normal_records = [record for record in records if record.car_scene == "无"]
    start_car_records = [record for record in records if record.car_scene == "开始用车"]
    end_car_records = [record for record in records if record.car_scene == "结束用车"]

    time_candidate_records = normal_records
    if len(normal_records) < 2:
        time_candidate_records = normal_records + start_car_records + end_car_records

    sign_in_record = min(time_candidate_records, key=lambda record: record.attendance_time, default=None)
    sign_out_record = max(time_candidate_records, key=lambda record: record.attendance_time, default=None)

    sign_in_time = sign_in_record.attendance_time if sign_in_record else None
    sign_out_time = sign_out_record.attendance_time if sign_out_record else None
    meets_8_hours = "否"
    if sign_in_time and sign_out_time and sign_out_time - sign_in_time >= timedelta(hours=8):
        meets_8_hours = "是"

    sign_in_plate, sign_out_plate, plate_consistency = choose_plate_result(start_car_records, end_car_records)
    remarks: list[str] = []
    if sign_in_record and sign_in_record.car_scene == "开始用车":
        remarks.append("签到时间使用开始用车时间")
    if sign_out_record and sign_out_record.car_scene == "结束用车":
        remarks.append("签退时间使用结束用车时间")
    if len(start_car_records) > 1:
        remarks.append(f"存在{count_text(len(start_car_records))}条开始用车")
    if len(end_car_records) > 1:
        remarks.append(f"存在{count_text(len(end_car_records))}条结束用车")

    return AuditResult(
        sign_in_time=sign_in_time,
        sign_out_time=sign_out_time,
        meets_8_hours=meets_8_hours,
        product_consistency=judge_product_consistency(person_format, sign_in_record, sign_out_record),
        sign_in_plate=sign_in_plate,
        sign_out_plate=sign_out_plate,
        plate_consistency=plate_consistency,
        remark="，".join(remarks),
    )


def first_sheet(workbook_path: Path) -> Worksheet:
    workbook = load_workbook(workbook_path, data_only=True)
    return workbook[workbook.sheetnames[0]]


def header_map(sheet: Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in sheet[1]:
        header = normalize_text(cell.value)
        if header and header not in result:
            result[header] = cell.column
    return result


def require_headers(headers: dict[str, int], required: Iterable[str], file_name: str) -> None:
    missing = [header for header in required if header not in headers]
    if missing:
        raise AttendanceAuditError(f"{file_name} 缺少必要列：{', '.join(missing)}")


def list_attendance_files(attendance_dir: str | Path) -> list[Path]:
    attendance_dir = Path(attendance_dir)
    if not attendance_dir.exists() or not attendance_dir.is_dir():
        raise AttendanceAuditError(f"打卡数据文件夹不存在：{attendance_dir}")

    files = sorted(
        path for path in attendance_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not files:
        raise AttendanceAuditError(f"打卡数据文件夹中没有可读取的 .xlsx 文件：{attendance_dir}")
    return files


def read_attendance_records(attendance_dir: str | Path) -> tuple[list[AttendanceRecord], int]:
    files = list_attendance_files(attendance_dir)
    records: list[AttendanceRecord] = []
    required = ["考勤类型", "姓名", "用户邮箱", "考勤时间", "主产品", "用车场景", "车牌号"]
    for file_path in files:
        sheet = first_sheet(file_path)
        headers = header_map(sheet)
        require_headers(headers, required, file_path.name)

        for row in range(2, sheet.max_row + 1):
            attendance_type = normalize_text(sheet.cell(row, headers["考勤类型"]).value)
            if attendance_type != "签到":
                continue
            time_value = parse_attendance_time(sheet.cell(row, headers["考勤时间"]).value)
            if time_value is None:
                continue
            name = normalize_text(sheet.cell(row, headers["姓名"]).value)
            if not name:
                continue
            records.append(
                AttendanceRecord(
                    name=name,
                    email=normalize_text(sheet.cell(row, headers["用户邮箱"]).value).lower(),
                    attendance_time=time_value,
                    product=normalize_text(sheet.cell(row, headers["主产品"]).value),
                    car_scene=normalize_text(sheet.cell(row, headers["用车场景"]).value),
                    plate_number=normalize_text(sheet.cell(row, headers["车牌号"]).value),
                    source_file=file_path.name,
                )
            )
    return records, len(files)


def copy_cell(source_cell: Any, target_cell: Any) -> None:
    target_cell.value = source_cell.value
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def merge_attendance_files(
    attendance_dir: str | Path,
    output_file: str | Path,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[Path, int, int]:
    files = list_attendance_files(attendance_dir)
    output_file = Path(output_file)
    progress = progress_callback or (lambda _message: None)

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "合并打卡数据"

    merged_row = 1
    data_row_count = 0
    expected_headers: list[str] | None = None

    for file_index, file_path in enumerate(files, start=1):
        progress(f"正在合并打卡数据：{file_index}/{len(files)} {file_path.name}")
        source_workbook = load_workbook(file_path, data_only=False)
        source_sheet = source_workbook[source_workbook.sheetnames[0]]
        headers = [normalize_text(source_sheet.cell(1, column).value) for column in range(1, source_sheet.max_column + 1)]
        if expected_headers is None:
            expected_headers = headers
            for column in range(1, source_sheet.max_column + 1):
                copy_cell(source_sheet.cell(1, column), output_sheet.cell(1, column))
            output_sheet.cell(1, source_sheet.max_column + 1).value = "来源文件"
            output_sheet.cell(1, source_sheet.max_column + 1).font = Font(bold=True)
            merged_row = 2
        elif headers != expected_headers:
            raise AttendanceAuditError(f"{file_path.name} 的表头与其他打卡数据文件不一致，无法合并")

        for source_row in range(2, source_sheet.max_row + 1):
            for column in range(1, source_sheet.max_column + 1):
                copy_cell(source_sheet.cell(source_row, column), output_sheet.cell(merged_row, column))
            output_sheet.cell(merged_row, source_sheet.max_column + 1).value = file_path.name
            merged_row += 1
            data_row_count += 1

    if expected_headers:
        for column in range(1, len(expected_headers) + 2):
            output_sheet.column_dimensions[output_sheet.cell(1, column).column_letter].width = 18
        output_sheet.freeze_panes = "A2"
        output_sheet.auto_filter.ref = f"A1:{output_sheet.cell(1, len(expected_headers) + 1).column_letter}{max(1, merged_row - 1)}"

    output_file.parent.mkdir(parents=True, exist_ok=True)
    try:
        output_workbook.save(output_file)
    except PermissionError as exc:
        raise AttendanceAuditError(f"无法保存合并打卡数据文件，可能已被 Excel 打开：{output_file}") from exc
    return output_file, len(files), data_row_count


def build_record_index(records: Iterable[AttendanceRecord]) -> tuple[dict[str, list[AttendanceRecord]], dict[tuple[str, str], list[AttendanceRecord]]]:
    by_name: dict[str, list[AttendanceRecord]] = defaultdict(list)
    by_name_email: dict[tuple[str, str], list[AttendanceRecord]] = defaultdict(list)
    for record in records:
        by_name[record.name].append(record)
        by_name_email[(record.name, record.email)].append(record)
    return by_name, by_name_email


def audit_personnel(
    person_file: str | Path,
    attendance_dir: str | Path,
    output_file: str | Path,
    progress_callback: Callable[[str], None] | None = None,
    merge_attendance: bool = False,
    merged_output_file: str | Path | None = None,
) -> AuditSummary:
    person_file = Path(person_file)
    output_file = Path(output_file)
    if not person_file.exists():
        raise AttendanceAuditError(f"人员信息表不存在：{person_file}")

    progress = progress_callback or (lambda _message: None)
    merged_output_path: Path | None = None
    if merge_attendance:
        if merged_output_file is None:
            merged_output_file = output_file.with_name(f"{output_file.stem}_合并打卡数据.xlsx")
        progress("正在合并打卡数据...")
        merged_output_path, _merged_file_count, _merged_row_count = merge_attendance_files(
            attendance_dir,
            merged_output_file,
            progress_callback=progress,
        )

    progress("正在读取打卡数据...")
    attendance_records, attendance_file_count = read_attendance_records(attendance_dir)
    by_name, by_name_email = build_record_index(attendance_records)

    progress("正在读取人员信息...")
    source_workbook = load_workbook(person_file)
    source_sheet = source_workbook[source_workbook.sheetnames[0]]
    if source_sheet.max_column < 6:
        raise AttendanceAuditError("人员信息表至少需要 A-F 六列")

    person_names = [
        normalize_text(source_sheet.cell(row, 2).value)
        for row in range(2, source_sheet.max_row + 1)
        if normalize_text(source_sheet.cell(row, 2).value)
    ]
    duplicate_names = sorted(name for name, count in Counter(person_names).items() if count > 1)

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = source_sheet.title

    for column in range(1, 7):
        if source_sheet.column_dimensions[source_sheet.cell(1, column).column_letter].width:
            output_sheet.column_dimensions[source_sheet.cell(1, column).column_letter].width = (
                source_sheet.column_dimensions[source_sheet.cell(1, column).column_letter].width
            )

    for row in range(1, source_sheet.max_row + 1):
        for column in range(1, 7):
            copy_cell(source_sheet.cell(row, column), output_sheet.cell(row, column))

    for index, header in enumerate(RESULT_HEADERS, start=7):
        cell = output_sheet.cell(1, index)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    for row in range(2, source_sheet.max_row + 1):
        name = normalize_text(source_sheet.cell(row, 2).value)
        email = normalize_text(source_sheet.cell(row, 3).value).lower()
        person_format = normalize_text(source_sheet.cell(row, 4).value)
        if not name:
            continue

        if name in duplicate_names:
            matched_records = by_name_email.get((name, email), [])
        else:
            matched_records = by_name.get(name, [])
        result = audit_person_records(person_format, matched_records)

        values = [
            result.sign_in_time,
            result.sign_out_time,
            result.meets_8_hours,
            result.product_consistency,
            result.sign_in_plate or "/",
            result.sign_out_plate or "/",
            result.plate_consistency,
            result.remark,
        ]
        for offset, value in enumerate(values, start=7):
            output_sheet.cell(row, offset).value = value
        output_sheet.cell(row, 7).number_format = "yyyy/mm/dd hh:mm:ss"
        output_sheet.cell(row, 8).number_format = "yyyy/mm/dd hh:mm:ss"

        if row % 20 == 0:
            progress(f"已核查 {row - 1} 人...")

    widths = {
        "G": 20,
        "H": 20,
        "I": 14,
        "J": 24,
        "K": 14,
        "L": 14,
        "M": 14,
        "N": 34,
    }
    for column, width in widths.items():
        output_sheet.column_dimensions[column].width = width
    output_sheet.freeze_panes = "A2"
    output_sheet.auto_filter.ref = f"A1:N{source_sheet.max_row}"

    output_file.parent.mkdir(parents=True, exist_ok=True)
    progress("正在保存结果...")
    try:
        output_workbook.save(output_file)
    except PermissionError as exc:
        raise AttendanceAuditError(f"无法保存输出文件，可能已被 Excel 打开：{output_file}") from exc
    progress(f"完成：{output_file}")

    return AuditSummary(
        person_count=max(0, source_sheet.max_row - 1),
        attendance_file_count=attendance_file_count,
        attendance_record_count=len(attendance_records),
        output_path=output_file,
        duplicate_names=duplicate_names,
        merged_output_path=merged_output_path,
    )
