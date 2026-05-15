from __future__ import annotations

import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from attendance_audit_core import (
    AttendanceRecord,
    audit_person_records,
    audit_personnel,
    judge_plate_consistency,
    judge_product_consistency,
    parse_attendance_time,
    normalize_project_id,
)


class AttendanceAuditCoreTests(unittest.TestCase):
    def record(
        self,
        name: str = "张三",
        email: str = "zhangsan@example.com",
        time: datetime = datetime(2026, 5, 13, 8, 0, 0),
        product: str = "5G NR",
        scene: str = "无",
        plate: str = "",
        attendance_type: str = "签到",
        site_code: str = "打卡站点-2025年新",
        project_id: str = "项目A",
        approval_status: str = "",
        range_start: datetime | None = None,
        range_end: datetime | None = None,
    ) -> AttendanceRecord:
        return AttendanceRecord(
            name=name,
            email=email,
            attendance_time=time,
            product=product,
            car_scene=scene,
            plate_number=plate,
            attendance_type=attendance_type,
            site_code=site_code,
            project_id=project_id,
            approval_status=approval_status,
            time_range_start=range_start or time,
            time_range_end=range_end or time,
        )

    def test_parse_attendance_time_accepts_string_and_datetime(self) -> None:
        value = datetime(2026, 5, 13, 8, 15, 31)
        self.assertEqual(parse_attendance_time(value), value)
        self.assertEqual(parse_attendance_time("2026/05/13 08:15:31"), value)
        self.assertEqual(
            parse_attendance_time("2026/01/23 18:00:00-2026/07/01 08:25:00"),
            datetime(2026, 7, 1, 8, 25, 0),
        )

    def test_eight_hour_boundary_is_yes(self) -> None:
        result = audit_person_records(
            "5G NR网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 8, 0, 0), product="5G NR"),
                self.record(time=datetime(2026, 5, 13, 16, 0, 0), product="5G NR"),
            ],
        )
        self.assertEqual(result.meets_8_hours, "是")
        self.assertEqual(result.product_consistency, "一致")
        self.assertEqual(result.remark, "")
        self.assertEqual(result.site_check, "符合")

    def test_project_id_normalization_removes_trailing_parenthesized_note(self) -> None:
        self.assertEqual(normalize_project_id("2023年四川移动技术服务项目(2人3车）"), "2023年四川移动技术服务项目")
        self.assertEqual(normalize_project_id("2023年四川移动技术服务项目（2人3车）"), "2023年四川移动技术服务项目")

    def test_project_check_uses_person_project_and_file_project(self) -> None:
        matching = audit_person_records(
            "LTE网络优化",
            [self.record(product="LTE", project_id="2023年四川移动技术服务项目")],
            person_project="2023年四川移动技术服务项目(2人3车）",
        )
        self.assertEqual(matching.project_check, "符合")

        mismatching = audit_person_records(
            "LTE网络优化",
            [self.record(product="LTE", project_id="2025年四川移动无线技术服务项目")],
            person_project="2023年四川移动技术服务项目(2人3车）",
        )
        self.assertEqual(mismatching.project_check, "不符")

    def test_site_check_fails_when_station_code_is_not_2025_new(self) -> None:
        result = audit_person_records(
            "5G NR网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 8, 0, 0), site_code="旧打卡站点"),
                self.record(time=datetime(2026, 5, 13, 17, 0, 0), site_code="打卡站点-2025年新"),
            ],
        )
        self.assertEqual(result.site_check, "不符")

    def test_site_check_passes_when_leave_range_covers_sign_date(self) -> None:
        result = audit_person_records(
            "5G NR网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 8, 0, 0), site_code="旧打卡站点"),
                self.record(time=datetime(2026, 5, 13, 17, 0, 0), site_code="旧打卡站点"),
                self.record(
                    time=datetime(2026, 5, 14, 18, 0, 0),
                    attendance_type="请假",
                    site_code="",
                    range_start=datetime(2026, 5, 12, 9, 0, 0),
                    range_end=datetime(2026, 5, 14, 18, 0, 0),
                ),
            ],
        )
        self.assertEqual(result.site_check, "符合")

    def test_site_check_passes_when_approved_makeup_covers_same_date(self) -> None:
        result = audit_person_records(
            "LTE网络优化",
            [
                self.record(time=datetime(2026, 5, 14, 8, 47, 40), product="LTE", site_code="旧站点"),
                self.record(time=datetime(2026, 5, 14, 17, 32, 47), product="LTE", site_code="另一个旧站点"),
                self.record(
                    time=datetime(2026, 5, 14, 18, 0, 0),
                    product="LTE",
                    attendance_type="补单",
                    site_code="打卡站点-成都电信枢纽中心-2025年新",
                    approval_status="审批通过",
                ),
            ],
        )
        self.assertEqual(result.site_check, "符合")

    def test_site_check_fails_when_makeup_is_not_valid(self) -> None:
        not_approved = audit_person_records(
            "LTE网络优化",
            [
                self.record(time=datetime(2026, 5, 14, 8, 47, 40), product="LTE", site_code="旧站点"),
                self.record(
                    time=datetime(2026, 5, 14, 18, 0, 0),
                    product="LTE",
                    attendance_type="补单",
                    site_code="打卡站点-成都电信枢纽中心-2025年新",
                    approval_status="审批中",
                ),
            ],
        )
        self.assertEqual(not_approved.site_check, "不符")

        wrong_date = audit_person_records(
            "LTE网络优化",
            [
                self.record(time=datetime(2026, 5, 14, 8, 47, 40), product="LTE", site_code="旧站点"),
                self.record(
                    time=datetime(2026, 5, 15, 18, 0, 0),
                    product="LTE",
                    attendance_type="补单",
                    site_code="打卡站点-成都电信枢纽中心-2025年新",
                    approval_status="审批通过",
                ),
            ],
        )
        self.assertEqual(wrong_date.site_check, "不符")

    def test_single_leave_record_outputs_leave_and_slashes(self) -> None:
        result = audit_person_records(
            "LTE网络优化",
            [
                self.record(
                    time=datetime(2026, 5, 15, 18, 0, 0),
                    product="LTE",
                    attendance_type="请假",
                    site_code="",
                    project_id="2023年四川移动无线技术服务项目",
                )
            ],
            person_project="2023年四川移动无线技术服务项目",
        )
        self.assertEqual(result.project_check, "请假")
        self.assertEqual(result.site_check, "请假")
        self.assertEqual(result.sign_in_time, "/")
        self.assertEqual(result.sign_out_time, "/")
        self.assertEqual(result.meets_8_hours, "/")
        self.assertEqual(result.product_consistency, "/")
        self.assertEqual(result.sign_in_plate, "/")
        self.assertEqual(result.sign_out_plate, "/")
        self.assertEqual(result.plate_consistency, "/")
        self.assertEqual(result.remark, "")

    def test_uses_car_scene_times_when_normal_records_less_than_two(self) -> None:
        result = audit_person_records(
            "5G NR网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 12, 0, 0), product="5G NR", scene="无"),
                self.record(time=datetime(2026, 5, 13, 8, 0, 0), product="5G NR", scene="开始用车", plate="川A12345"),
                self.record(time=datetime(2026, 5, 13, 18, 0, 0), product="5G NR", scene="结束用车", plate="川A12345"),
            ],
        )
        self.assertEqual(result.sign_in_time, datetime(2026, 5, 13, 8, 0, 0))
        self.assertEqual(result.sign_out_time, datetime(2026, 5, 13, 18, 0, 0))
        self.assertEqual(result.product_consistency, "一致")
        self.assertEqual(result.plate_consistency, "一致")
        self.assertEqual(result.remark, "签到时间使用开始用车时间，签退时间使用结束用车时间")

    def test_multiple_start_car_records_match_any_end_plate(self) -> None:
        result = audit_person_records(
            "LTE网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 8, 18, 57), product="LTE", scene="无"),
                self.record(time=datetime(2026, 5, 13, 8, 56, 58), product="LTE", scene="开始用车", plate="川ACX5359"),
                self.record(time=datetime(2026, 5, 13, 10, 40, 8), product="LTE", scene="开始用车", plate="川AC75307"),
                self.record(time=datetime(2026, 5, 13, 15, 50, 19), product="LTE", scene="结束用车", plate="川AC75307"),
            ],
        )
        self.assertEqual(result.sign_in_plate, "川AC75307")
        self.assertEqual(result.sign_out_plate, "川AC75307")
        self.assertEqual(result.plate_consistency, "一致")
        self.assertEqual(result.remark, "签退时间使用结束用车时间，存在两条开始用车")

    def test_normal_records_keep_priority_when_two_or_more_exist(self) -> None:
        result = audit_person_records(
            "LTE网络优化",
            [
                self.record(time=datetime(2026, 5, 13, 9, 0, 0), product="LTE", scene="无"),
                self.record(time=datetime(2026, 5, 13, 17, 0, 0), product="LTE", scene="无"),
                self.record(time=datetime(2026, 5, 13, 8, 0, 0), product="5G NR", scene="开始用车", plate="川A12345"),
                self.record(time=datetime(2026, 5, 13, 18, 0, 0), product="5G NR", scene="结束用车", plate="川A12345"),
            ],
        )
        self.assertEqual(result.sign_in_time, datetime(2026, 5, 13, 9, 0, 0))
        self.assertEqual(result.sign_out_time, datetime(2026, 5, 13, 17, 0, 0))
        self.assertEqual(result.product_consistency, "一致")
        self.assertEqual(result.remark, "")

    def test_product_consistency_all_cases(self) -> None:
        good = self.record(product="LTE")
        bad = self.record(product="5G NR")

        self.assertEqual(judge_product_consistency("LTE网络优化", good, good), "一致")
        self.assertEqual(judge_product_consistency("LTE网络优化", bad, good), "签到制式不一致")
        self.assertEqual(judge_product_consistency("LTE网络优化", good, bad), "签退制式不一致")
        self.assertEqual(judge_product_consistency("LTE网络优化", bad, bad), "签到和签退制式不一致")

    def test_plate_consistency_all_cases(self) -> None:
        self.assertEqual(judge_plate_consistency("川A12345", "川A12345"), "一致")
        self.assertEqual(judge_plate_consistency("川A12345", "川B12345"), "不一致")
        self.assertEqual(judge_plate_consistency("", "川B12345"), "缺少签到")
        self.assertEqual(judge_plate_consistency("川A12345", ""), "缺少签退")
        self.assertEqual(judge_plate_consistency("", ""), "未用车")

    def test_duplicate_name_uses_name_and_email_when_exporting(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            person_file = root / "人员信息.xlsx"
            attendance_dir = root / "打卡数据导出"
            output_file = root / "结果.xlsx"
            attendance_dir.mkdir()

            person_wb = Workbook()
            person_ws = person_wb.active
            person_ws.append(["项目号", "姓名", "用户邮箱", "打卡制式", "工号（电旗）", "姓名（电旗）"])
            person_ws.append(["P1", "王杰", "first@example.com", "5G NR网络优化", "001", "王杰.1"])
            person_ws.append(["P2", "王杰", "second@example.com", "LTE网络优化", "002", "王杰.2"])
            person_ws.append(["P3", "李四", "lisi@example.com", "LTE网络优化", "003", "李四"])
            person_wb.save(person_file)

            attendance_wb = Workbook()
            attendance_ws = attendance_wb.active
            attendance_ws.title = "data"
            attendance_ws.append(["考勤类型", "姓名", "用户邮箱", "考勤时间", "站点编码", "审批状态", "主产品", "用车场景", "车牌号"])
            attendance_ws.append(["签到", "王杰", "first@example.com", "2026/05/13 08:00:00", "打卡站点-2025年新", "", "5G NR", "无", ""])
            attendance_ws.append(["签到", "王杰", "first@example.com", "2026/05/13 17:00:00", "打卡站点-2025年新", "", "5G NR", "无", ""])
            attendance_ws.append(["签到", "王杰", "second@example.com", "2026/05/13 09:00:00", "打卡站点-2025年新", "", "LTE", "无", ""])
            attendance_ws.append(["签到", "王杰", "second@example.com", "2026/05/13 16:00:00", "打卡站点-2025年新", "", "LTE", "无", ""])
            attendance_ws.append(["签到", "李四", "wrong@example.com", "2026/05/13 08:30:00", "打卡站点-2025年新", "", "LTE", "无", ""])
            attendance_ws.append(["签到", "李四", "another@example.com", "2026/05/13 18:00:00", "打卡站点-2025年新", "", "LTE", "无", ""])
            attendance_ws.append(["签到", "李四", "another@example.com", "2026/05/13 07:50:00", "打卡站点-2025年新", "", "LTE", "开始用车", "川A12345"])
            attendance_ws.append(["签到", "李四", "another@example.com", "2026/05/13 18:05:00", "打卡站点-2025年新", "", "LTE", "结束用车", "川A12345"])
            attendance_ws.append(["请假", "李四", "another@example.com", "2026/01/23 18:00:00-2026/07/01 08:25:00", "", "", "LTE", "无", ""])
            attendance_wb.save(attendance_dir / "P1.xlsx")

            summary = audit_personnel(person_file, attendance_dir, output_file)
            self.assertEqual(summary.person_count, 3)
            self.assertEqual(summary.attendance_file_count, 1)
            self.assertEqual(summary.attendance_record_count, 8)
            self.assertEqual(summary.duplicate_names, ["王杰"])
            self.assertIsNone(summary.merged_output_path)

            output_wb = load_workbook(output_file, data_only=True)
            output_ws = output_wb.active

            self.assertEqual(output_ws.cell(2, 7).value, "符合")
            self.assertEqual(output_ws.cell(2, 8).value, "符合")
            self.assertEqual(output_ws.cell(2, 9).value, datetime(2026, 5, 13, 8, 0, 0))
            self.assertEqual(output_ws.cell(2, 10).value, datetime(2026, 5, 13, 17, 0, 0))
            self.assertEqual(output_ws.cell(2, 11).value, "是")
            self.assertEqual(output_ws.cell(2, 12).value, "一致")
            self.assertEqual(output_ws.cell(2, 13).value, "/")
            self.assertEqual(output_ws.cell(2, 14).value, "/")
            self.assertEqual(output_ws.cell(2, 15).value, "未用车")
            self.assertEqual(output_ws.cell(2, 16).value, None)

            self.assertEqual(output_ws.cell(3, 9).value, datetime(2026, 5, 13, 9, 0, 0))
            self.assertEqual(output_ws.cell(3, 10).value, datetime(2026, 5, 13, 16, 0, 0))
            self.assertEqual(output_ws.cell(3, 11).value, "否")
            self.assertEqual(output_ws.cell(3, 12).value, "一致")

            self.assertEqual(output_ws.cell(4, 9).value, datetime(2026, 5, 13, 8, 30, 0))
            self.assertEqual(output_ws.cell(4, 10).value, datetime(2026, 5, 13, 18, 0, 0))
            self.assertEqual(output_ws.cell(4, 13).value, "川A12345")
            self.assertEqual(output_ws.cell(4, 14).value, "川A12345")
            self.assertEqual(output_ws.cell(4, 15).value, "一致")

    def test_merge_attendance_files_when_requested(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            person_file = root / "人员信息.xlsx"
            attendance_dir = root / "打卡数据导出"
            output_file = root / "结果.xlsx"
            merged_file = root / "合并打卡数据.xlsx"
            attendance_dir.mkdir()

            person_wb = Workbook()
            person_ws = person_wb.active
            person_ws.append(["项目号", "姓名", "用户邮箱", "打卡制式", "工号（电旗）", "姓名（电旗）"])
            person_ws.append(["P1", "张三", "zhangsan@example.com", "LTE网络优化", "001", "张三"])
            person_wb.save(person_file)

            for index in range(1, 3):
                attendance_wb = Workbook()
                attendance_ws = attendance_wb.active
                attendance_ws.title = "data"
                attendance_ws.append(["考勤类型", "姓名", "用户邮箱", "考勤时间", "站点编码", "审批状态", "主产品", "用车场景", "车牌号"])
                attendance_ws.append(["签到", "张三", "zhangsan@example.com", f"2026/05/13 0{index}:00:00", "打卡站点-2025年新", "", "LTE", "无", ""])
                attendance_ws.append(["出差", "张三", "zhangsan@example.com", "2026/01/23 18:00:00-2026/07/01 08:25:00", "", "", "LTE", "无", ""])
                attendance_wb.save(attendance_dir / f"SIGN-{index}.xlsx")

            summary = audit_personnel(
                person_file,
                attendance_dir,
                output_file,
                merge_attendance=True,
                merged_output_file=merged_file,
            )

            self.assertEqual(summary.attendance_file_count, 2)
            self.assertEqual(summary.attendance_record_count, 2)
            self.assertEqual(summary.merged_output_path, merged_file)

            merged_wb = load_workbook(merged_file, data_only=True)
            merged_ws = merged_wb.active
            self.assertEqual(merged_ws.max_row, 5)
            self.assertEqual(merged_ws.max_column, 11)
            self.assertEqual(merged_ws.cell(1, 1).value, "项目号")
            self.assertEqual(merged_ws.cell(1, 11).value, "来源文件")
            self.assertEqual(merged_ws.cell(2, 1).value, "SIGN-1")
            self.assertEqual(merged_ws.cell(2, 2).value, "签到")
            self.assertEqual(merged_ws.cell(3, 2).value, "出差")


if __name__ == "__main__":
    unittest.main()
